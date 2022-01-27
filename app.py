import mariadb
import re
import hashlib
import random, string
import openpyxl as xl
import pandas as pd
from datetime import date, datetime
from flask import *
from flask_cors import CORS
from functools import wraps
import ipaddress

class Database:
    def __init__(self, **dns):
        self.dns = dns
        self.dbh = None

    def _open(self):
        self.dbh = mariadb.connect(**self.dns)

    def _close(self):
        self.dbh.close()

    def _commit(self):
        self.dbh.commit()

    def query(self, stmt, *args, **kwargs):
        #print(id(self))
        self._open()
        print( 'STMT is ' + stmt )
        if kwargs.get('prepared', False):
            cursor = self.dbh.cursor(prepared=True, dictionary=True, buffered=True)
            cursor.execute(stmt, args)
        elif kwargs.get('duplicated', False):
            cursor = self.dbh.cursor(buffered=True)
            cursor.execute(stmt)
            fields = [md[0] for md in cursor.description]
        else:
            cursor = self.dbh.cursor(dictionary=True, buffered=True)
            cursor.execute(stmt)
        if 'INSERT' in stmt or 'UPDATE' in stmt or 'DELETE' in stmt:
            self._commit()
            data = None
        else:
            data = cursor.fetchall()
        if kwargs.get('last_id', False):
            cursor.execute( 'SELECT last_insert_id();' )
            inserted_id = cursor.fetchone()['last_insert_id()']
            cursor.close()
            self._close()
            #print( 'Inserted ID is ' + str( inserted_id  ))
            return inserted_id
        elif kwargs.get('duplicated', False):
            return [fields, data]
        else:
            cursor.close()
            self._close()
            #print( 'Data is ' + str( data ))
            return data

dns = {
        'user': 'tomoki',
        'host': 'localhost',
        'password': 'utq0975e',
        'database': 'abldb'
        }

#dns = {
#        'user': 'abldb',
#        'host': 'localhost',
#        'password': 'CC#x-#hW/p?R@SMe',
#        'database': 'abldb'
#        }

#root_dir = ('/abldb-api')
root_dir = ('/api')

def random_string(n):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=n))

def get_cursor():
    connection = mariadb.connect(
            user = 'root',
            password = 'utq0975e',
            database = 'abldb',
            host = '127.0.0.1'
            )
    return connection.cursor(dictionary=True)

def up_matched(m):
    s = m.group(1)
    return s.upper()

def snake_to_camel(string):
    pattern = re.compile(r'_([a-z])')
    result = pattern.sub(up_matched, string)
    return result

def camel_to_snake(string):
    pattern = re.compile(r'([A-Z])')
    result = pattern.sub(r'_\1', string)
    return result.lower()

def format_dic_keys(tuple_of_dic, direction):
    new_list = []
    for row in tuple_of_dic:
        new_row = format_dic( row, direction )
        new_list.append( new_row )
    return new_list

def format_dic(dic, direction):
    new_row = {}
    for k, v in dic.items():
        if direction == 'camel':
            new_key = snake_to_camel(k)
        elif direction == 'snake':
            new_key = camel_to_snake(k)
        else:
            raise ValueError("error dayo")
        new_row[new_key] = v
    return new_row

def validiate_token(token):
    if token == 'hanamogera':
        return 0
    else:
        return None

def format_value(value):
    if value is None:
        return "NULL"
    elif isinstance( value, int ):
        return str( value )
    elif isinstance( value, date ):
        return "\'" + value.strptime( "%Y-%m-%d" ) + "\'"
    elif value == True:
        return "\'b\'1"
    elif  value == False:
        return "\'b\'0"
    else:
        return f"\'{str( value )}\'"

def format_data_to_insert(data):
    data = format_dic(data, 'snake')
    data_pairs = map( lambda k: '`' + k + '`' + " = " + format_value( data[k] ), data )
    return ','.join( data_pairs )

def format_data_to_cast(data):
    data  = format_dic( data, 'camel' )
    #print( data )
    data_to_cast = {}
    for k, v in data.items():
        if v == b'\x00':
            data_to_cast[k] = False
        elif v == b'\x01':
            data_to_cast[k] = True
        elif isinstance( v, date ):
            data_to_cast[k] = v.strftime( '%Y-%m-%d' )
        else:
            data_to_cast[k] = v
    return data_to_cast

def create_new_ucg():
    db = Database(**dns)
    query = '''
        INSERT INTO `ucg` () VALUES();
    '''
    new_id = db.query( query, last_id=True )
    return new_id

def create_new_medicine():
    db = Database(**dns)
    query = '''
        INSERT INTO `internal_medicine` () VALUES ();
    '''
    new_id = db.query( query, last_id=True )
    return new_id

def create_new_blood():
    db = Database(**dns)
    query = '''
        INSERT INTO `blood_exam` () VALUES ();
    '''
    new_id = db.query( query, last_id=True )
    return new_id

def create_new_holter(follow_up_id):
    db = Database(**dns)
    query = f'''
        INSERT INTO `holter` ( `follow_up_id` ) VALUES ( { follow_up_id } );
    '''

def delete_following_ablation(abl_id):
    db = Database(**dns)
    query = '''
        DELETE FROM `following_ablation` WHERE `following_ablation_id` = { abl_id };
    '''

def check_token(token):
    db = Database(**dns)
    query = f'''
        SELECT * FROM logins
            WHERE `token` = '{ token }'
                AND `is_logged_out` = b'0';
    '''
    result = db.query( query )
    if result == []:
        return False
    else:
        return result[0][ 'user_id' ]

app = Flask(__name__)
CORS(app)

def test1(f):
    def test2(*args, **kwargs):
        print( 'this is test 1' )
        return f(*args, *kwargs)
    return test2

def token_gate(func):
    @wraps( func )
    def query_token( *args, **kwargs ):
        token_header = request.headers.get( "Authorization" )
        if token_header == None:
            return jsonify({ 'error': 'Token invalid' }), 403
        token = token_header.split()[1]
        logined_user = check_token( token )
        kwargs[ 'logined_user_id' ] = logined_user
        if logined_user:
            return func( *args, **kwargs )
        else:
            return jsonify({ 'error': 'Token invalid' }), 403
    return query_token



@app.route(root_dir)
@test1
def hello_world():
    return '<html><body><p>It works.</p></body></html>'

@app.route(root_dir + '/login', methods=['POST'])
def authenticate_user():
    db = Database(**dns)
    login_request = request.json
    if login_request['order'] == 'salt':
        query = f'''
            SELECT `salt` FROM `users`
                WHERE `user_id` = "{ login_request['user'] }";
        '''
        result = db.query( query )
        if len(result) == 0:
            return jsonify({'salt': random_string(64) }), 200
        else:
            return jsonify({'salt': result[0]['salt']}), 200
    elif login_request['order'] == 'auth':
        user_id = login_request['user']
        user_salt = login_request['userSalt']
        challenge_hash = login_request['challengeHash']
        query = f'''
            SELECT cast(`password_hash` as char) AS password_hash FROM `users`
                WHERE `user_id` = "{ user_id }";
        '''
        result = db.query( query )
        if len(result) == 0:
            return 'Invalid', 404
        else:
            password_hash = result[0]['password_hash']
            hash_seed = password_hash + user_salt
            response_hash = hashlib.sha256(hash_seed.encode('UTF-8')).hexdigest()
            if challenge_hash == response_hash:
                token = random_string(64)
                #print( token )
                query = f'''
                    INSERT INTO `logins`
                        ( `user_id`, `token` )
                    VALUES ( '{user_id}', '{token}' );
                '''
                db.query( query )
                query_hospital_id = f'''
                    SELECT `hospital_id` FROM `users`
                        WHERE `user_id` = '{ user_id }';
                    '''
                hospital_id = db.query( query_hospital_id )[0]['hospital_id']
                query_hospital_name = f'''
                    SELECT `hospital_name` FROM `hospital`
                        WHERE hospital_id = { hospital_id }
                '''
                hospital_name = db.query( query_hospital_name )[0][ 'hospital_name' ]
                return jsonify({
                    'user_id': user_id,
                    'hospital': hospital_name,
                    'token': token
                })
            else:
                return 'Invalid', 404
    elif login_request['order'] == 'token':
        query = f'''
            SELECT * FROM logins
                WHERE `token` = '{logins_request['token']}';
        '''
        result = db.query( query )[0]
        if result is None:
            return 'Invalid', 404
        else:
            return 'ok', 200
    elif login_request['order'] == 'logout':
        token = login_request['token']
        if check_token( token ):
            query = f'''
                UPDATE `logins`
                    SET `is_logged_out` = b'1'
                WHERE `token` = '{token}';
            '''
            return 'ok', 200
        else:
            return 'Invalid', 404




@app.route(root_dir + '/patientlist/<int:page>', methods=['GET'])
@token_gate
def list_patients(page, **kwargs):
    min_number = ( page * 100 ) - 100
    max_number = page * 100
    db = Database(**dns)
    query_hospital = f'''
        SELECT `hospital_id` FROM `users`
            WHERE `user_id` = '{ kwargs[ 'logined_user_id' ] }';
    '''
    hospital_id = db.query( query_hospital )[0]['hospital_id']
    query_patient_count = f'''
        SELECT MAX(`patient_number`) AS last_number FROM `patient_list`
            WHERE `hospital_id` = { hospital_id };
    '''
    patients_count = db.query( query_patient_count )[0]['last_number']
    query = f'''
        SELECT * FROM patient_list
            WHERE `hospital_id` = { hospital_id }
                AND `patient_number` > { min_number }
                AND `patient_number` <= { max_number }
            ORDER BY `patient_number` DESC;
    '''
    patients_list = db.query( query )
    for row in patients_list:
        pt_number = row['patient_serial_number']
        query_follow_abl = f'''
            SELECT `following_ablation_id` FROM `following_ablation`
                WHERE `patient_serial_number` = { pt_number }
                ORDER BY `date`;
        '''
        res_follow_abl = db.query( query_follow_abl )
        row['following_ablations'] = list( map( lambda x: x['following_ablation_id'], res_follow_abl ))
    patients_list = list( map( format_data_to_cast, patients_list ) )
    data = { "patients": patients_list, "totalPatients": patients_count  }
    return jsonify( data )

@app.route(root_dir + '/patients/<int:patient_serial_number>', methods=['GET'])
@token_gate
def give_a_patient(patient_serial_number, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT * FROM `patient_list`
            WHERE `patient_serial_number` = { patient_serial_number };
    '''
    patient = db.query( query )[0]
    query_follow_abl = f'''
        SELECT `following_ablation_id` FROM `following_ablation`
            WHERE `patient_serial_number` = { patient_serial_number }
            ORDER BY `date`;
    '''
    res_follow_abl = db.query( query_follow_abl )
    patient['following_ablations'] = list( map( lambda x: x['following_ablation_id'], res_follow_abl ))
    return jsonify( format_data_to_cast( patient ) )

@app.route(root_dir + '/baseline/new', methods=['GET'])
@token_gate
def get_new_pt_number(**kwargs):
    db = Database(**dns)
    header = request.headers.get("Authorization")
    _,token = header.split()
    user_id = check_token(token)
    query_hospital = f'''
        SELECT `hospital_id` FROM `users`
            WHERE `user_id` = '{ user_id }';
    '''
    hospital_id = db.query( query_hospital )[0][ 'hospital_id' ]
    query_pt_number = f'''
        SELECT MAX( `patient_number` ) AS `patient_number` FROM `patients`
            WHERE `hospital_id` = { hospital_id };
    '''
    last_pt_number = db.query( query_pt_number )[0][ 'patient_number' ]
    if last_pt_number is None:
        last_pt_number = 0
    query = f'''
        INSERT INTO `patients` 
            (`registered_by`, `hospital_id`, `patient_number`)
        VALUES ('{user_id}', {hospital_id}, {last_pt_number + 1});
    '''
    patient_id = db.query( query, last_id=True )
    ucg_id = create_new_ucg()
    query = f'''
        UPDATE `patients`
            SET `ucg_id` = { ucg_id }
        WHERE `patient_serial_number` = { patient_id };
    '''
    db.query( query )
    return str( patient_id )

@app.route(root_dir + '/baseline/<int:patient_serial_number>', methods=['GET'])
@token_gate
def give_baseline_data(patient_serial_number, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT * FROM `patients`
            WHERE `patient_serial_number` = {patient_serial_number};
    '''
    baseline = db.query( query )[0]
    return jsonify( format_data_to_cast( baseline ))

@app.route(root_dir + '/baseline/<int:patient_serial_number>', methods=['POST'])
@token_gate
def update_baseline_data(patient_serial_number, **kwargs):
    db = Database(**dns)
    baseline_data = format_data_to_insert( request.json )
    query = f'''
        UPDATE `patients` 
            SET { baseline_data }
        WHERE `patient_serial_number` = {patient_serial_number};
    '''
    #print( query )
    db.query( query )
    return "ok"

@app.route(root_dir + '/ucg/new', methods=['GET'])
@token_gate
def get_new_ucg_number(**kwargs):
    db = Database(**dns)
    query = '''
        INSERT INTO `ucg`
        (`lvdd`)
        VALUES (NULL);
    '''
    new_id = db.query( query, last_id=True )
    #print( new_id )
    return jsonify( new_id )

@app.route(root_dir + '/ucg/<int:ucg_id>', methods=['GET'])
@token_gate
def give_ucg_data(ucg_id, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT * FROM `ucg`
            WHERE `ucg_id` = { ucg_id };
    '''
    ucg = db.query( query )[0]
    return jsonify( format_data_to_cast( ucg ) )

@app.route(root_dir + '/ucg/<int:ucg_id>', methods=['POST'])
@token_gate
def update_ucg_data(ucg_id, **kwargs):
    db = Database(**dns)
    ucg_data = format_data_to_insert( request.json )
    query = f'''
        UPDATE `ucg`
            SET { ucg_data }
        WHERE `ucg_id` = { ucg_id };
    '''
    db.query( query )
    return "ok"

@app.route(root_dir + '/holter/<int:holter_id>', methods=['GET'])
@token_gate
def give_holter_data(holter_id, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT * FROM `holter`
            WHERE `holter_id` = { holter_id };
    '''
    data = db.query( query )[0]
    return jsonify( format_data_to_cast( data ) )

@app.route(root_dir + '/holter/<int:holter_id>', methods=['POST'])
@token_gate
def update_holter_data(holter_id, **kwargs):
    db = Database(**dns)
    holter_data = format_data_to_insert( request.json )
    query = f'''
        UPDATE `holter`
            SET { holter_data }
        WHERE `holter_id` = { holter_id };
    '''
    db.query( query )
    return "ok"

@app.route(root_dir + '/blood/<int:blood_id>', methods=['GET'])
@token_gate
def give_blood_data(blood_id, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT * FROM `blood_exam`
            WHERE `blood_id` = { blood_id };
    '''
    data = db.query( query )[0]
    return jsonify( format_data_to_cast( data ) )

@app.route(root_dir + '/blood/<int:blood_id>', methods=['POST'])
@token_gate
def update_blood_data(blood_id, **kwargs):
    db = Database(**dns)
    blood_data = format_data_to_insert( request.json )
    query = f'''
        UPDATE `blood_exam`
            SET { blood_data }
        WHERE `blood_id` = { blood_id };
    '''
    db.query( query )
    return "ok"

@app.route(root_dir + '/holter_list/<int:ptSerial>', methods=['GET'])
@token_gate
def give_holter_list(ptSerial, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT `holter_id`, `date` FROM `holter`
            WHERE `patient_serial_number` = { ptSerial };
    '''
    result = db.query( query )
    if len( result ) == 0:
        data = []
    else:
        print( result )
        data = result

    return jsonify( data )


@app.route(root_dir + '/holter/<int:ptSerial>/new', methods=['GET'])
@token_gate
def create_new_holter(ptSerial, **kwargs):
    db = Database(**dns)
    query = f'''
        INSERT INTO `holter`
            ( `patient_serial_number` )
        VALUES ( { ptSerial } );
    '''
    db.query( query )
    return give_holter_list(ptSerial)


@app.route(root_dir + '/1st-abl/<int:patient_serial_number>', methods=['GET'])
@token_gate
def give_first_abl_data(patient_serial_number, **kwargs):
    db = Database(**dns)
    q_data = f'''
        SELECT * FROM `first_ablation`
            WHERE `patient_serial_number` = { patient_serial_number };
    '''
    q_new_data = f'''
        INSERT INTO `first_ablation`
            (`patient_serial_number`)
        VALUES ( { patient_serial_number } );
    '''
    result = db.query( q_data )
    if len( result ) == 0:
        db.query( q_new_data )
        abl_data = db.query(q_data)[0]
    else:
        abl_data = result[0]
    #print( abl_data )
    return jsonify( format_data_to_cast( abl_data ) )

@app.route(root_dir + '/1st-abl/<int:patient_serial_number>', methods=['POST'])
@token_gate
def update_first_abl_data(patient_serial_number, **kwargs):
    db = Database(**dns)
    abl_data = format_data_to_insert( request.json )
    query = f'''
        UPDATE `first_ablation`
            SET { abl_data }
        WHERE `patient_serial_number` = { patient_serial_number };
    '''
    db.query( query )
    return 'First ablation update: SUCCESS'

@app.route(root_dir + '/1st-abl/<int:first_abl_id>/medication_id', methods=['GET'])
@token_gate
def give_med_id_for_first_abl(first_abl_id, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT `internal_medicine_id` FROM `first_ablation`
            WHERE `first_ablation_id` = { first_abl_id };
    '''
    q_new_medicine = "INSERT INTO `internal_medicine` () VALUES ();"
    result = db.query( query )[0]
    medicine_id = result[ 'internal_medicine_id' ]
    if medicine_id is None:
        medicine_id = db.query( q_new_medicine, last_id=True )
        q_register_medicine = f'''
            UPDATE `first_ablation`
                SET `internal_medicine_id` = { medicine_id }
                WHERE `first_ablation_id` = { first_abl_id };
        '''
        db.query( q_register_medicine )
    return jsonify(medicine_id)

@app.route(root_dir + '/medication/<int:medication_id>', methods=['GET'])
@token_gate
def give_medication_data(medication_id, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT * FROM `internal_medicine`
            WHERE `internal_medicine_id` = { medication_id }
    '''
    medication_data = db.query( query )[0]
    return jsonify( format_data_to_cast( medication_data ) )

@app.route(root_dir + '/medication/<int:medication_id>', methods=['POST'])
@token_gate
def update_medication_data(medication_id, **kwargs):
    db = Database(**dns)
    medication_data = format_data_to_insert( request.json )
    query = f'''
        UPDATE `internal_medicine`
            SET { medication_data }
        WHERE `internal_medicine_id` = { medication_id };
    '''
    db.query( query )
    return "ok"

@app.route(root_dir + '/following_ablation/new/<int:patient_serial_number>', methods=['GET'])
@token_gate
def get_new_follow_ablation_number(patient_serial_number, **kwargs):
    db = Database(**dns)
    query = f'''
        INSERT INTO `following_ablation`
            ( patient_serial_number ) VALUES ( { patient_serial_number } );
    '''
    follow_ablation_id = db.query( query, last_id=True )
    ucg_id = create_new_ucg()
    medicine_id = create_new_medicine()
    query = f'''
        UPDATE `following_ablation`
            SET `ucg_id` = { ucg_id }, `internal_medicine_id` = { medicine_id }
        WHERE `following_ablation_id` =  { follow_ablation_id };
    '''
    db.query( query )
    return jsonify( follow_ablation_id )

@app.route(root_dir + '/following_ablation/<int:following_ablation_id>', methods=['GET'])
@token_gate
def give_following_ablation_data(following_ablation_id, **kwargs):
    db = Database(**dns)
    query = f'''
        SELECT * FROM `following_ablation`
            WHERE `following_ablation_id` = { following_ablation_id };
    '''
    follow_ablation = db.query( query )[0]
    return jsonify( format_data_to_cast( follow_ablation ) )

@app.route(root_dir + '/following_ablation/<int:following_ablation_id>', methods=['POST'])
@token_gate
def update_following_ablation_data(following_ablation_id, **kwargs):
    db = Database(**dns)
    following_ablation_data = format_data_to_insert( request.json )
    if request.json.get('order') == 'delete':
        query = f'''
            DELETE FROM `following_ablation` WHERE `following_ablation_id` = { following_ablation_id };
        '''
    else:
        query = f'''
            UPDATE `following_ablation`
                SET { following_ablation_data }
            WHERE `following_ablation_id` = { following_ablation_id };
        '''
    db.query( query )
    return "ok"

@app.route(root_dir + '/followup/<int:patient_serial_number>', methods=['GET'])
@token_gate
def give_followup_data(patient_serial_number, **kwargs):
    db = Database(**dns)
    q_data = f'''
        SELECT * FROM `follow_up`
            WHERE `patient_serial_number` = { patient_serial_number };
    '''
    q_new_data = f'''
        INSERT INTO `follow_up`
            (`patient_serial_number`)
        VALUES ( { patient_serial_number } );
    '''
    result = db.query( q_data )
    if len( result ) == 0:
        follow_up_id = db.query( q_new_data, last_id=True )
        ucg_id1 = create_new_ucg()
        ucg_id2 = create_new_ucg()
        ucg_id3 = create_new_ucg()
        blood_id1 = create_new_blood()
        blood_id2 = create_new_blood()
        blood_id3 = create_new_blood()

        q_add_sections = f'''
            UPDATE `follow_up`
                SET `ucg_id1` = { ucg_id1 }, `ucg_id2` = { ucg_id2 }, `ucg_id3` = { ucg_id3 },
                    `blood_id1` = { blood_id1 }, `blood_id2` = { blood_id2 }, `blood_id3` = { blood_id3 }
            WHERE `follow_up_id` = { follow_up_id };
        '''

        db.query( q_add_sections )

        follow_up_data = db.query(q_data)[0]
    else:
        follow_up_data = result[0]

        for i in range(1,4):
            if not follow_up_data[ 'ucg_id' + str(i) ]:
                ucg_id = create_new_ucg()
                q_add_ucg = f'''
                    UPDATE `follow_up`
                        SET `ucg_id{ str(i) }` = { ucg_id }
                        WHERE `follow_up_id` = { follow_up_data[ 'follow_up_id' ] };
                '''
                db.query( q_add_ucg )
                follow_up_data[ 'ucg_id' + str(i) ] = ucg_id

            if not follow_up_data[ 'blood_id' + str(i) ]:
                blood_id = create_new_blood()
                q_add_blood = f'''
                    UPDATE `follow_up`
                        SET `blood_id{ str(i) }` = { blood_id }
                        WHERE `follow_up_id` = { follow_up_data[ 'follow_up_id' ] };
                '''
                db.query( q_add_ucg )
                follow_up_data[ 'blood_id' + str(i) ] = blood_id

    return jsonify( format_data_to_cast( follow_up_data ) )

@app.route(root_dir + '/followup/<int:patient_serial_number>', methods=['POST'])
@token_gate
def update_followup_data(patient_serial_number, **kwargs):
    db = Database(**dns)
    follow_up_data = format_data_to_insert( request.json )
    query = f'''
        UPDATE `follow_up`
            SET { follow_up_data }
        WHERE `patient_serial_number` = { patient_serial_number };
    '''
    db.query( query )
    return 'Follow up update: SUCCESS'

@app.route( root_dir + '/existing_users', methods=['GET'] )
@token_gate
def give_existing_users( **kwargs ):
    db = Database(**dns)
    q_users = f'''
        SELECT `user_id` FROM `users`;
    '''
    res_users = db.query( q_users )
    users = list( map( lambda x: x[ 'user_id' ], res_users ) )

    q_institutes = f'''
        SELECT * FROM `hospital`;
    '''
    res_institutes = db.query( q_institutes )
    ids = list( map( lambda x: x[ 'hospital_id' ], res_institutes ) )
    hospitals = list( map( lambda x: x[ 'hospital_name' ], res_institutes ) )
    institutes = dict( zip( ids, hospitals ) )

    return jsonify( { 'ids': users, 'institutes': institutes } )

@app.route( root_dir + '/new_user', methods=['POST'] )
@token_gate
def add_new_user( **kwargs ):
    db = Database( **dns )

    new_user = request.json
    user_id = new_user[ 'user' ]
    hospital_id = int( new_user[ 'institute' ] )
    password_hash = new_user[ 'hash' ]
    salt = new_user[ 'salt' ]

    q_newuser = f'''
        INSERT INTO `users` ( `user_id`, `hospital_id`, `password_hash`, `salt` )
            VALUES ( '{ user_id }', { hospital_id }, CAST( '{ password_hash }' AS BINARY ), '{ salt }' );
    '''

    try:
        db.query( q_newuser )
    except:
        return jsonify( { 'message': 'データベースエラーです' } )

    return jsonify( { 'message': '登録しました' } )

@app.route( root_dir + '/new_institute', methods=['POST'] )
@token_gate
def add_new_institute( **kwargs ):
    db = Database( **dns )
    new_institute = request.json[ 'institute' ];


    q_newinstitute = f'''
        INSERT INTO `hospital` ( `hospital_name` )
            VALUES ( '{ new_institute }' );
    '''

    print( q_newinstitute )

    try:
        institute_num = db.query( q_newinstitute, last_id=True )
    except:
        return jsonify( { 'message': 'データベースエラーです' } )

    return jsonify( {
        'message': 'OK',
        'instituteNum': institute_num
    } )

@app.route( root_dir + '/new_password', methods=['POST'] )
@token_gate
def change_password( **kwargs ):
    db = Database( **dns )
    request_data = request.json

    user_salt = request_data['userSalt']
    challenge_hash = request_data['challengeHash']
    new_salt = request_data['newSalt']
    new_hash = request_data['newHash']
    user_id = request_data['user']

    query_validate_hash = f'''
        SELECT cast(`password_hash` as char) AS password_hash FROM `users`
            WHERE `user_id` = '{ user_id }';
    '''
    result = db.query( query_validate_hash )

    if len(result) == 0:
        return jsonify({ 'message': 'データベースエラーです' })
    else:
        existing_hash = result[0]['password_hash']
        hash_seed = existing_hash + user_salt
        response_hash = hashlib.sha256(hash_seed.encode('UTF-8')).hexdigest()
        if challenge_hash == response_hash:
            query_password = f'''
                UPDATE `users`
                    SET `password_hash` = CAST( '{ new_hash }' AS BINARY ),
                        `salt` = '{ new_salt }'
                    WHERE `user_id` = '{ user_id }';
            '''
            
            try:
                db.query( query_password )
            except:
                return jsonify({ 'message': 'データベースエラーです' })

            return jsonify({ 'message': 'OK' })
        else:
            return jsonify({ 'message': 'パスワードが間違っています' })

@app.route( root_dir + '/export_xlsx', methods=['GET'] )
def make_excel_file():
    remote_addr = ipaddress.ip_address(request.remote_addr)
    if remote_addr != ipaddress.ip_address("127.0.0.1"):
        return "Access Denied", 403

    hospital_id = request.args.get("hpid")
    exporting_path = request.args.get("path")
    #print(exporting_path)
    db = Database( **dns )

    def get_columns( result ):
        #q_columns = f'''
        #    SHOW COLUMNS FROM `{ table }`;
        #'''
        #res_columns = db.query( q_columns )
        #columns = list( map( lambda x: x[ 'Field' ], res_columns ) )
        #return columns
        return result[0].keys()

    def table_to_excel( table, ws ):
        if isinstance( table[0], dict ):
            columns = table[0].keys()
        elif isinstance( table[0], list ):
            columns = table[0]
            if isinstance(table[1][0], tuple):
                table = table[1]
            else:
                table.pop(0)

        x = 1
        for column in columns:
            ws.cell( row=1, column=x, value=column)
            x += 1

        for i, table_row in enumerate( table ):
            for j, column in enumerate( columns ):
                value = None
                if isinstance( table[0], dict ):
                    value = table_row[ column ]
                elif isinstance( table[0], tuple ):
                    value = table_row[ j ]
                elif isinstance( table[0], list ):
                    try:
                        value = table_row[ j ]
                    except IndexError:
                        value = None
                if value == b'\x00':
                    if column == 'sex':
                        value = 'Female'
                    else:
                        value = False
                elif value == b'\x01':
                    if column == 'sex':
                        value = 'Male'
                    else:
                        value = True
                #print( value )
                ws.cell( row=2+i, column=j+1, value=value )

    # baseline
    q_baseline = f'''
        SELECT * FROM `patients`
            LEFT JOIN `ucg`
                ON `patients`.`ucg_id` = `ucg`.`ucg_id`
            WHERE `patients`.`hospital_id` = { hospital_id };
    '''
    baseline_table = db.query( q_baseline )

    wb = xl.Workbook()
    baseline_ws = wb.active
    baseline_ws.title = "Baseline"
    table_to_excel( baseline_table, baseline_ws )
    for count,col in enumerate([1,2,3,5]):
        baseline_ws.delete_cols(col - count)

    #1st session
    q_first_session = f'''
        SELECT `patients`.`patient_number`, `first_ablation`.* FROM `patients`
            LEFT JOIN `first_ablation`
                ON `patients`.`patient_serial_number` = `first_ablation`.`patient_serial_number`
            WHERE `patients`.`hospital_id` = { hospital_id };
    '''
    first_session_table = db.query( q_first_session )

    first_session_ws = wb.create_sheet( title="1st session" )
    table_to_excel( first_session_table, first_session_ws )
 
    for count,col in enumerate([2,4,5]):
        first_session_ws.delete_cols(col - count)

    #1st session medicine
    q_first_medicine = f'''
        SELECT `patients`.`patient_number`, `internal_medicine`.* FROM `patients`
            LEFT JOIN `first_ablation`
                ON `patients`.`patient_serial_number` = `first_ablation`.`patient_serial_number`
            LEFT JOIN `internal_medicine`
                ON `first_ablation`.`internal_medicine_id` = `internal_medicine`.`internal_medicine_id`
            WHERE `patients`.`hospital_id` = { hospital_id };
    '''
    first_medicine_table = db.query( q_first_medicine )

    first_medicine_ws = wb.create_sheet( title="1st session内服薬" )
    table_to_excel( first_medicine_table, first_medicine_ws )

    for count,col in enumerate([2]):
        first_medicine_ws.delete_cols(col - count)

    #Following session
    q_following_session = f'''
        SELECT  `patients`.`patient_number`, `following_ablation`.*, `ucg`.*, `internal_medicine`.* FROM `patients`
            LEFT JOIN `following_ablation`
                ON `following_ablation`.`patient_serial_number` = `patients`.`patient_serial_number`
            LEFT JOIN `ucg`
                ON `following_ablation`.`ucg_id` = `ucg`.`ucg_id`
            LEFT JOIN `internal_medicine`
                ON `following_ablation`.`internal_medicine_id` = `internal_medicine`.`internal_medicine_id`
            WHERE `patients`.`hospital_id` = { hospital_id }
            ORDER BY `patient_number`, `following_ablation`.`date`;
    '''
    following_session_table = db.query( q_following_session )

    # horizonize
    following_session_df = pd.DataFrame( following_session_table )
    following_session_df = following_session_df.drop(['following_ablation_id', 'patient_serial_number', 'ucg_id', 'internal_medicine_id'], axis=1)
    patients = following_session_df[ "patient_number" ].unique()

    fs_table = []
    iterated_count = 0
    for patient in patients:
        following_session_a_patient = following_session_df[ following_session_df[ "patient_number" ] == patient ]
        
        patient_row = []
        for count, row in enumerate( following_session_a_patient.itertuples() ):
            row_list = list( row )
            row_list.pop(0)
            patient_row.extend( row_list )
            if count > iterated_count:
                iterated_count = count
        
        #print( patient_row )
        fs_table.append( patient_row )

    original_headers = following_session_df.columns.tolist()
    #original_headers.pop(0)
    header_list = []
    for i in range(iterated_count):
        header_list.extend(original_headers)

    fs_table.insert(0,header_list)


    #print( fs_table )


    following_session_ws = wb.create_sheet( title="following session" )
    table_to_excel( fs_table, following_session_ws )
    #for count,col in enumerate([2,4,5]):
    #    following_session_ws.delete_cols(col - count)

    #Follow up
    q_followup = f'''
        SELECT `patients`.`patient_number`, `follow_up`.*,
                `ucg1`.*, `ucg2`.*, `ucg3`.*, `blood1`.*, `blood2`.*, `blood3`.* FROM `patients` 
            LEFT JOIN `follow_up` ON `patients`.`patient_serial_number` = `follow_up`.`patient_serial_number`
            LEFT JOIN `ucg` as `ucg1` ON `follow_up`.`ucg_id1` = `ucg1`.`ucg_id`
            LEFT JOIN `ucg` as `ucg2` ON `follow_up`.`ucg_id2` = `ucg2`.`ucg_id`
            LEFT JOIN `ucg` as `ucg3` ON `follow_up`.`ucg_id3` = `ucg3`.`ucg_id`
            LEFT JOIN `blood_exam` as `blood1` ON `follow_up`.`blood_id1` = `blood1`.`blood_id`
            LEFT JOIN `blood_exam` as `blood2` ON `follow_up`.`blood_id2` = `blood2`.`blood_id`
            LEFT JOIN `blood_exam` as `blood3` ON `follow_up`.`blood_id3` = `blood3`.`blood_id`
            WHERE `patients`.`hospital_id` = { hospital_id };
    '''
    followup_table = db.query( q_followup, duplicated=True )
    #print(followup_table[0])
    #print(followup_table[1])
    followup_ws = wb.create_sheet( title="follow up" )
    table_to_excel( followup_table, followup_ws )

    wb.save( exporting_path )

    return( "dekitayo!" )


if __name__ == '__main__':
    app.run( threaded=True, debug=True )
